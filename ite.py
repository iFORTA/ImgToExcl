import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
from openpyxl import Workbook
import re
import os
import io
import base64
import numpy as np

def open_img():
    global file_path
    global suggest_width ,suggest_height
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.ppm *.pgm *.tif *.tiff")])
    if file_path:
        gen_pixlsize = 200
        file_path_name.set(file_path)
        image = Image.open(file_path)
        im_size = image.size
        imgH.set(im_size[0])
        imgW.set(im_size[1])
        # 显示图片预览
        if im_size[0] > im_size[0]:
            suggest_width = int(im_size[0] / (im_size[0] / gen_pixlsize))
            suggest_height = int(im_size[1] / (im_size[0] / gen_pixlsize))
        else:
            suggest_width = int(im_size[0] / (im_size[1] / gen_pixlsize))
            suggest_height = int(im_size[1] / (im_size[1] / gen_pixlsize))
        rImg=image.resize((suggest_width,suggest_height))
        update_image_preview(rImg)

def update_image_preview(image):
    img_preview = ImageTk.PhotoImage(image)
    image_preview_label.config(image=img_preview)
    image_preview_label.image = img_preview

def exp_excel():
    
    if file_path:
        opt_value = var_option.get()
        if opt_value == 3:
            custom_width = TxtImgWidth.get()
            custom_height = TxtImgHeight.get()
            if not custom_width or not custom_height:
                tk.messagebox.showwarning("警告", "请在自定义宽和高中输入值。")
                return
            width = int(TxtImgWidth.get())
            height = int(TxtImgHeight.get())
        elif opt_value == 2:
            width = int(imgW.get())
            height = int(imgH.get())
        elif opt_value == 1:
            width = suggest_width
            height = suggest_height
        image = Image.open(file_path)
        im_rsize = image.resize((width, height))
        img = np.array(im_rsize)
        s = np.shape(img)
        wb = Workbook()
        ws = wb.create_sheet('myphotos', 0)
        for i in range(0, s[0]):
            for j in range(0, s[1]):
                ws.cell(row=i + 1, column=j + 1).value = str(img[i][j][0]) + '_' + str(img[i][j][1]) + '_' + str(
                    img[i][j][2])
        filename = os.path.basename(file_path)
        pattern = r'^(.*?)\.[^.]*$'
        match = re.match(pattern, filename)
        if match:
            result = match.group(1)
            lname = result + '.xlsx'
            wb.save(lname)
            tk.messagebox.showinfo("成功", "导出成功。")
        else:
            tk.messagebox.showwarning("错误", "导出失败。")
            return


def toggle_code():
    global text_visible
    if text_visible:
        # 如果文本框可见，隐藏文本框
        code_text.pack_forget()
        text_visible = False
    else:
        # 如果文本框不可见，显示文本框
        code_text.delete("1.0", tk.END)
        code = """
Sub FillCellsWithRGBColorsAndHideValues()
    Dim ws As Worksheet
    Dim cell As Range
    Dim rgbString As String
    Dim rgbArray() As String
    Dim r As Integer, g As Integer, b As Integer
    
    ' 指定要处理的工作表
    Set ws = ThisWorkbook.Sheets("myphotos") ' 替换为你的工作表名称
    
    ' 禁用屏幕刷新以提高性能
    Application.ScreenUpdating = False
    
    ' 循环遍历工作表中的所有单元格
    For Each cell In ws.UsedRange
        ' 获取单元格的文本
        rgbString = cell.Value
        
        ' 分割RGB字符串
        rgbArray() = Split(rgbString, "_")
        
        ' 确保RGB字符串包含3个值
        If UBound(rgbArray) = 2 Then
            r = CInt(rgbArray(0))
            g = CInt(rgbArray(1))
            b = CInt(rgbArray(2))
            
            ' 设置单元格的背景色为解析出的RGB值
            cell.Interior.Color = RGB(r, g, b)
            
            ' 将单元格的文本值设置为空
            cell.Value = ""
        End If
    Next cell
    
    ' 启用屏幕刷新
    Application.ScreenUpdating = True
End Sub

' Excel单元格设置为正方形效果最佳
' 20像素行高:12
' 20像素行高:1.44

"""
        code_text.insert(tk.END, code)
        code_text.pack()
        text_visible = True




text_visible = False

icon_base64 = """
AAABAAEAGBgAAAEAIACICQAAFgAAACgAAAAYAAAAMAAAAAEAIAAAAAAAAAkAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACidIAAnnx8AJ54fACeeH
wAnnh8AJ54fACeeHwAonSAAJ54fACefHwAAAAAAAAAAAAAAAAAnnh8AJ54fACeeHwAnnh8AJ54fACeeHw
Annh8AJ54fACeeHwAnnh8AJ54fACieIAAnnh8AJ54fACafHgEnnh8KJ54fDCeeHwMnnh8AJ54fACeeHwA
nnh8AAAAAAAAAAAAnnh8AJ54fACadHgAonyABKJ8gASifIAEonyABKJ8gASifIAEonyABKJ8gASqgIgAn
nh8NJ54fUieeH5wnnh/BJ54fxSeeH6gnnh9jJ54fFSeeHwAnnh8AJ54fAAAAAAAnnh8AJ54fKieeH5Ynn
h+hJ54foCeeH6Annh+gJ54foCeeH6Annh+gJ54fnyeeH58nnh+8J54f9yeeH/8nnh//J54f/yeeH/8nnh/9J54
fwSeeHzcnnh8AJp4eACefHwAnnh8AJ54fdieeH7snnh9bJ54fWyeeH1snnh9bJ54fWyeeH1snnh9aJ54fcye
eH+wnnh//J54f/yeeH/4nnh/7J54f/yeeH/8nnh//J54f/yeeH9Unnh8tJ54fACeeHwAnnh8AJ54ffCeeH4In
nh8AJp8eByafHgomnx4KJp8eCiafHgomnh4IJ54ffCeeH/0nnh//J54f8ieeH40nnh/HJ54f/yeeH/8nnh//J
54f/yeeH/8nnh+mJ54fByeeHwAnnh8AJ54feyeeH4Mnnh8FJ54fcieeH5Innh+RJ54fkSeeH5Annh+bJ54f6ye
eH/8nnh/aJ54fVyeeHwknnh+uJ54f/yeeH/8nnh//J54f/yeeH/8nnh/tJ54fNCeeHwAnnh8AJ54feyeeH4Mnn
h8AJ54fIyeeHy4nnh8tJ54fLSeeHysnnh9hJ54f9SeeH7Qnnh8uJ54fACedHwEnnh8qJ54fYyeeH74nnh/9J54f/ye
eH/8nnh//J54fYieeHwAnnh8AJ54feyeeH4Mnnh8FJ54fcSeeH5Mnnh+RJ54fkSeeH48nnh+xJ54f9yeeH3on
nh8FJ54fACedHwAJlwEAJ54fACeeHw4nnh+NJ54f/CeeH/8nnh//J54fdCeeHwAnnh8AJ54feyeeH4MoniAAJ
54fKCeeHzUnnh80J54fNCeeHzEnnh9mJ54f+CeeH/gnnh+aJ54fHCeeHwEnnh99J54fnieeH2Mnnh8nJ54foi
eeH/8nnh//J54fYyeeHwAnnh8AJ54feyeeH4Mnnh8FJ54fbCeeH4onnh+JJ54fiSeeH4knnh+TJ54f5ie
eH/8nnh//J54fxSeeH0Qnnh+8J54f/yeeH/8nnh/JJ54fdieeH+Ynnh/vJ54fNSeeHwAnnh8AJ54fe
yeeH4Mmnh8AJ54fHieeHygnnh8oJ54fKCeeHxUnnh8QJ54fpyeeH/8nnh//J54f/yeeH+cnnh/qJ5
4f/yeeH/8nnh//J54f5yeeH+gnnh+qJ54fCCeeHwAnnh8AJ54feyeeH4MonSAGJ54frieeH+gn
nh/kJ54f6CeeH3Innh8lJ54fiSeeH8wnnh//J54f/yeeH/8nnh//J54f/yeeH/8nnh//J54f/yeeH9gnn
h8xJ54fACeeHwAnnh8AJ54feyeeH4MonSAHJ54fwyeeH/8nnh//J54f/yeeH4Innh8OJ54fOCeeH2Annh/gJ
54f/yeeH/8nnh//J54f/yeeH/8nnh//J54f6yeeH0Ynnh8AJp8eACefHwAnnh8AJ54feyeeH4MonSAHJ54fw
ieeH/8nnh//J54f/yeeH4Annh8iJ54feCeeH3knnh+CJ54fiieeH6Unnh/RJ54f7SeeH64nnh9/J54fvieeH
xsnnh8AJ54fAAAAAAAnnh8AJ54feyeeH4MonSAHJ54fxCeeH/8nnh//J54f/yeeH4Innh8UJ54fSyeeH00nn
h9MJ54fSieeHxgnnh87J54fsyeeHw4nnh8XJ54ftCeeHx0nnh8AAAAAAAAAAAAnnh8AJ54feyeeH4IjoxsAJ
54fVSeeH3Innh9xJ54fcieeHzcnnh8fJ54fayeeH24nnh9uJ54fayeeHx0nnh8uJ54frieeHwonnh8aJ54ft
CeeHx0nnh8AAAAAAAAAAAAnnh8AJ54ffCeeH5onnh8cJ54fGyeeHxsnnh8bJ54fGyeeHx0nnh8eJ54fGyeeHxsnn
h8bJ54fGyeeHxsnnh9QJ54ftSeeHwonnh8aJ54ftCeeHx0nnh8AAAAAAAAAAAAnnh8AJ54fSieeH8snnh/PJ5
4fzyeeH78nnh+/J54fvyeeH78nnh+/J54fvyeeH78nnh+/J54fvyeeH78nnh/KJ54fhyeeHwEnnh8bJ54ftCe
eHx0nnh8AAAAAAAAAAAAnnh8AJ54fAyeeHxcnnh+QJ54fkieeHxwnnh8fJ54fHyeeHx8nnh8fJ54fHyeeHx
8nnh8fJ54fHyeeHx8nnh8dJ54fBieeHwAnnh8YJ54ftCeeHx0nnh8AAAAAAAAAAAAnnh8AJ54fACeeHwAnn
h97J54fpieeHzEnnh8yJ54fMieeHzInnh8yJ54fMieeHzInnh8yJ54fMieeHzInnh8yJ54fNCeeHzMnnh9R
J54fwieeHxwnnh8AAAAAAAAAAAAAAAAAJ54fACeeHwAnnh8+J54fuSeeH7onnh+5J54fuSeeH7knnh+5J54
fuSeeH7knnh+5J54fuSeeH7knnh+5J54fuSeeH7knnh++J54fiSeeHwgnnh8AAAAAAAAAAAAAAAAAJ54fACe
eHwAloB0AJ54fCyeeHw8nnh8PJ54fDyeeHw8nnh8PJ54fDyeeHw8nnh8PJ54fDyeeHw8nnh8PJ54fDyeeHw
8nnh8PJ54fBCeeHwAnnh8AAAAAAAAAAAAAAAAAAQMBACeeHwAnnh8AJ54fACeeHwAnnh8AJ54fACeeHwAnn
h8AJ54fACeeHwAnnh8AJ54fACeeHwAnnh8AJ54fACeeHwAnnh8AJ54fACeeHwAnnh8AAAAAAAAAAAD/4Ac
AAAADAAAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMA
AAADAAAAAwAAAAMAAAADAAAAAwCAAAMAgAADAIAAAwA=
"""

icon_bytes = base64.b64decode(icon_base64)

icon_image = Image.open(io.BytesIO(icon_bytes))



root = tk.Tk()
tk_icon = ImageTk.PhotoImage(icon_image)
root.iconphoto(True, tk_icon)

root.title("像素数据导出")


file_path_name = tk.StringVar()
imgH = tk.IntVar()
imgW = tk.IntVar()
var_option = tk.IntVar()
var_option.set(1)
file_path = ''






frame1 = tk.Frame(root)
frame1.pack(padx=10, pady=10)

frame2 = tk.Frame(root)
frame2.pack(padx=10, pady=10)



frame6 = tk.Frame(root)
frame6.pack()


frame3 = tk.Frame(root)
frame3.pack(padx=10, pady=10)

frame4 = tk.Frame(root)
frame4.pack(padx=10, pady=10)

frame5 = tk.Frame(root)
frame5.pack(padx=10, pady=10)

frame8 = tk.Frame(root)
frame8.pack(padx=10, pady=10)



frame7 = tk.Frame(root)
frame7.pack(padx=10, pady=10)



toggle_button = tk.Button(frame5, text="VBA代码示例", command=toggle_code)
toggle_button.pack(side=tk.LEFT)

code_text = tk.Text(frame8)
code_text.pack()






image_preview_label = tk.Label(frame6)
image_preview_label.pack()


tk.Label(frame1, text="选择文件").pack(side=tk.LEFT)

TxtOpenImg = tk.Entry(frame1, textvariable=file_path_name, width=40)
TxtOpenImg.pack(side=tk.LEFT)

BtnOpenImg = tk.Button(frame1, text="打开图片", command=open_img)
BtnOpenImg.pack(side=tk.LEFT)

tk.Label(frame2, text="原始图片大小").pack(side=tk.LEFT)
tk.Label(frame2, text="宽:").pack(side=tk.LEFT)
TxtOrgImgW = tk.Entry(frame2, textvariable=imgW, width=10)
TxtOrgImgW.pack(side=tk.LEFT)

tk.Label(frame2, text="高:").pack(side=tk.LEFT)
TxtOrgImgH = tk.Entry(frame2, textvariable=imgH, width=10)
TxtOrgImgH.pack(side=tk.LEFT)

tk.Label(frame3, text="修改图片大小").pack(side=tk.LEFT)
LabImgWidth = tk.Label(frame3, text="宽:")
LabImgWidth.pack(side=tk.LEFT)
TxtImgWidth = tk.Entry(frame3, width=10)
TxtImgWidth.pack(side=tk.LEFT)

LabImgHeight = tk.Label(frame3, text="高:")
LabImgHeight.pack(side=tk.LEFT)
TxtImgHeight = tk.Entry(frame3, width=10)
TxtImgHeight.pack(side=tk.LEFT)

rb_option1 = tk.Radiobutton(frame4, text="推荐比例", variable=var_option, value=1)
rb_option2 = tk.Radiobutton(frame4, text="原始比例", variable=var_option, value=2)
rb_option3 = tk.Radiobutton(frame4, text="自定比例", variable=var_option, value=3)

rb_option1.pack(side=tk.LEFT)
rb_option2.pack(side=tk.LEFT)
rb_option3.pack(side=tk.LEFT)

BtnExpExcl = tk.Button(frame5, text="EXCEL导出", command=exp_excel)
BtnExpExcl.pack(side=tk.LEFT)

tk.Label(frame7, text="@iFORT ").pack(side=tk.LEFT)

tk.Label(frame7, text=" v1.2309.2").pack(side=tk.LEFT)
root.mainloop()


# pyinstaller --onefile --noconsole --icon=C:\Users\LauXing\Pictures\icon\aefgl.ico  ite.py